"""Command line entry point.

`pyproject.toml` has declared `crucible = "crucible.cli:app"` since the first commit, but
this module never existed, so `uv run crucible` failed with an ImportError while
`crucible-app` worked. Fixed here rather than by deleting the script line, because the
pipeline now needs a headless path: the deliverable is a file, and generating a file
should not require starting a web server.

Commands are added as their stage lands. A command that cannot yet do its job is absent
rather than stubbed - a stub that accepts an invocation and quietly does nothing is the
same class of failure as a verifier that abstains without saying so.
"""

from __future__ import annotations

import logging
from pathlib import Path

import typer
from rich.console import Console
from rich.table import Table

from crucible.emit.columns import N_COLUMNS, ColumnError, load_reference_header, validate_header
from crucible.ingest import IngestError, read_products

app = typer.Typer(
    add_completion=False,
    no_args_is_help=True,
    help="Certified product data enrichment for industrial catalogs.",
)
console = Console()


@app.callback()
def _main(verbose: bool = typer.Option(False, "--verbose", "-v", help="Log at DEBUG.")) -> None:
    logging.basicConfig(
        level=logging.DEBUG if verbose else logging.INFO,
        format="%(asctime)s %(message)s",
        datefmt="%H:%M:%S",
    )


@app.command("verify-format")
def verify_format(
    path: Path = typer.Argument(..., help="A delivery CSV or XLSX to check."),
) -> None:
    """Assert a file carries the 252 delivery columns, exactly and in order.

    Judge-facing: the brief forbids removing, renaming or reordering any header, and a
    file that violates that fails before its data quality is ever assessed.
    """
    try:
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True)
            sheet = workbook.active
            header = tuple(str(c.value or "") for c in next(sheet.iter_rows(max_row=1)))
            workbook.close()
        else:
            header = load_reference_header(path)
        validate_header(header)
    except (ColumnError, FileNotFoundError) as exc:
        console.print(f"[bold red]FAIL[/] {path.name}: {exc}")
        raise typer.Exit(1) from exc

    console.print(f"[bold green]OK[/] {path.name}: {N_COLUMNS} columns, exact order")


@app.command()
def inspect(
    input: Path = typer.Option(..., "--input", "-i", help="A distributor CSV export."),
    limit: int | None = typer.Option(None, "--limit", "-n", help="Read only the first N rows."),
) -> None:
    """Summarise what an input export actually contains.

    Written because the first useful question about an unseen evaluation file is not
    "what did we extract" but "how much signal is even present" - how many rows carry a
    real brand rather than a placeholder, and how long the descriptions are.
    """
    try:
        products = read_products(input, limit)
    except IngestError as exc:
        console.print(f"[bold red]FAIL[/] {exc}")
        raise typer.Exit(1) from exc

    lengths = sorted(len(p.description) for p in products)
    branded = sum(1 for p in products if p.brand)
    with_mpn_in_desc = sum(
        1 for p in products if p.mpn and p.mpn.casefold() in p.description.casefold()
    )

    table = Table(title=f"{input.name}", show_header=False, box=None)
    table.add_row("products", f"{len(products):,}")
    table.add_row("distinct SKUs", f"{len({p.sku for p in products}):,}")
    table.add_row(
        "with a real brand",
        f"{branded:,} ({branded / len(products):.0%}) - the rest are placeholders",
    )
    table.add_row(
        "part number echoed in description",
        f"{with_mpn_in_desc:,} ({with_mpn_in_desc / len(products):.0%})"
        " - the identity check's second channel",
    )
    table.add_row(
        "description length",
        f"min {lengths[0]}, median {lengths[len(lengths) // 2]}, max {lengths[-1]} chars",
    )
    console.print(table)


@app.command()
def route(
    input: Path = typer.Option(..., "--input", "-i", help="A distributor CSV export."),
    limit: int | None = typer.Option(None, "--limit", "-n", help="Read only the first N rows."),
    show: int = typer.Option(0, "--show", help="Print this many unclassified rows."),
) -> None:
    """Classify an export and report where the products landed.

    Coverage is reported rather than optimised for. The unclassified tail is printed on
    request because it is the honest output of the step: those rows still export, with
    the four classification columns left blank, and seeing them is how the taxonomy gets
    extended deliberately instead of by guesswork.
    """
    from crucible.route import GENERIC_CATEGORY_ID, CascadeRouter

    try:
        products = read_products(input, limit)
    except IngestError as exc:
        console.print(f"[bold red]FAIL[/] {exc}")
        raise typer.Exit(1) from exc

    router = CascadeRouter()
    routings = router.route_all(products)

    table = Table(title=f"routing {input.name}")
    table.add_column("category")
    table.add_column("rows", justify="right")
    table.add_column("share", justify="right")
    for category, count in sorted(router.stats.by_category.items(), key=lambda kv: -kv[1]):
        style = "dim" if category == GENERIC_CATEGORY_ID else ""
        table.add_row(category, f"{count:,}", f"{count / len(products):.1%}", style=style)
    console.print(table)
    console.print(
        f"{router.stats.coverage:.1%} classified; "
        f"{router.stats.ambiguous} rows had no clear winner and abstained"
    )

    if show:
        console.print("\n[bold]unclassified[/] (these export with blank Dept/Class/Fine):")
        shown = 0
        for product_, routing in zip(products, routings, strict=True):
            if routing.category_id == GENERIC_CATEGORY_ID and shown < show:
                console.print(f"  [{routing.method}] {product_.description}")
                shown += 1


def main() -> None:
    app()


if __name__ == "__main__":
    main()


@app.command()
def enrich(
    input_path: Path = typer.Option(..., "--input", "-i", help="Input CSV of sparse products."),
    out: Path = typer.Option(Path("runs/demo"), "--out", "-o", help="Directory for outputs."),
    limit: int | None = typer.Option(None, "--limit", "-n", help="Only the first N products."),
    fill_mode: str = typer.Option("certified", "--fill-mode", help="certified | grounded | all."),
    threshold: float | None = typer.Option(
        None, "--threshold", help="Nonconformity cutoff for certified mode."
    ),
    model: str | None = typer.Option(None, "--model", help="Ollama model id."),
    no_llm: bool = typer.Option(False, "--no-llm", help="Rules only; no inference."),
    no_preflight: bool = typer.Option(
        False, "--no-preflight", help="Skip the Ollama/GPU readiness check."
    ),
    concurrency: int = typer.Option(
        4, "--concurrency", "-j", min=1, help="Products in flight at once."
    ),
) -> None:
    """Turn a sparse input CSV into the 252-column delivery file.

    Writes delivery.csv, delivery.xlsx and evidence.csv into the output directory.
    """
    from crucible.emit.rows import EmitPolicy, FillMode
    from crucible.emit.writer import write_csv, write_evidence, write_xlsx
    from crucible.enrich import coverage_by_column
    from crucible.enrich import enrich as run_enrich
    from crucible.preflight import PreflightError, check_ollama

    if not no_llm and not no_preflight:
        try:
            console.print(f"[dim]preflight: {check_ollama(model)}[/dim]")
        except PreflightError as exc:
            console.print(f"[red]{exc}[/red]")
            raise typer.Exit(2) from exc

    try:
        mode = FillMode(fill_mode)
    except ValueError:
        console.print(
            f"[red]unknown --fill-mode {fill_mode!r}; "
            f"expected one of {', '.join(m.value for m in FillMode)}[/red]"
        )
        raise typer.Exit(2) from None

    policy = EmitPolicy(fill_mode=mode, threshold=threshold)

    if mode is FillMode.CERTIFIED and threshold is None:
        console.print(
            "[yellow]certified mode with no --threshold: nothing has been calibrated, so "
            "no attribute values will be published. Pass --threshold, or use "
            "--fill-mode grounded to export every grounded value.[/yellow]"
        )

    with console.status("[bold]enriching...") as status:

        def progress(index: int, total: int, sku: str) -> None:
            status.update(f"[bold]{index}/{total}[/bold]  {sku}")

        try:
            result = run_enrich(
                input_path,
                limit=limit,
                policy=policy,
                model=model,
                use_llm=not no_llm,
                progress=progress,
                concurrency=concurrency,
            )
        except IngestError as exc:
            console.print(f"[red]{exc}[/red]")
            raise typer.Exit(2) from exc

    out.mkdir(parents=True, exist_ok=True)
    n_csv = write_csv(result.rows, out / "delivery.csv")
    n_ev = write_evidence(result.evidence_pairs(), out / "evidence.csv")
    try:
        write_xlsx(result.rows, out / "delivery.xlsx")
        xlsx_note = str(out / "delivery.xlsx")
    except RuntimeError as exc:
        xlsx_note = f"skipped ({exc})"

    console.print(f"\n[green]{result.stats.summary()}[/green]")

    table = Table(title="Outputs", show_header=False)
    table.add_row("delivery.csv", f"{n_csv} rows -> {out / 'delivery.csv'}")
    table.add_row("delivery.xlsx", xlsx_note)
    table.add_row("evidence.csv", f"{n_ev} cells -> {out / 'evidence.csv'}")
    console.print(table)

    coverage = coverage_by_column(result.rows)
    top = Table(title="Most-populated columns")
    top.add_column("Column")
    top.add_column("Rows", justify="right")
    top.add_column("%", justify="right")
    for column, count in list(coverage.items())[:15]:
        top.add_row(column, str(count), f"{100 * count / max(result.stats.products, 1):.0f}%")
    console.print(top)


@app.command()
def evaluate(
    truth: Path = typer.Option(
        Path("Unihack_ Expected Output - Delivery Format.csv"),
        "--truth",
        "-t",
        help="A delivery sheet containing enriched rows to score against.",
    ),
    input_path: Path | None = typer.Option(
        None, "--input", "-i", help="Catalog to measure compliance over (default: the truth rows)."
    ),
    limit: int | None = typer.Option(None, "--limit", "-n", help="Only the first N products."),
    model: str | None = typer.Option(None, "--model", help="Ollama model id."),
) -> None:
    """Score the pipeline on the three metrics the client's guide names.

    Field-level accuracy against labelled rows, character-limit compliance, and
    controlled-vocabulary compliance.
    """
    import csv as _csv
    import tempfile

    from crucible.enrich import enrich as run_enrich
    from crucible.evaluate import (
        check_limits,
        check_vocabulary,
        compare_rows,
        format_report,
        load_truth,
        truth_as_input,
    )
    from crucible.ontology import GENERIC_CATEGORY_ID, generic_schema, load_all
    from crucible.preflight import PreflightError, check_ollama

    try:
        console.print(f"[dim]preflight: {check_ollama(model)}[/dim]")
    except PreflightError as exc:
        console.print(f"[red]{exc}[/red]")
        raise typer.Exit(2) from exc

    try:
        labelled = load_truth(truth)
    except (OSError, ValueError) as exc:
        console.print(f"[red]{exc}[/red]")
        raise typer.Exit(2) from exc

    console.print(f"[dim]{len(labelled)} labelled row(s) in {truth.name}[/dim]")

    # Re-run our own pipeline over the same inputs the labelled rows came from.
    with tempfile.TemporaryDirectory() as tmp:
        replay = Path(tmp) / "truth-input.csv"
        rows = truth_as_input(labelled)
        with replay.open("w", newline="", encoding="utf-8") as handle:
            writer = _csv.DictWriter(handle, fieldnames=list(rows[0]))
            writer.writeheader()
            writer.writerows(rows)

        with console.status("[bold]scoring against the labelled rows..."):
            scored = run_enrich(replay, policy=None, model=model)

    accuracy = compare_rows(scored.rows, labelled)

    # Compliance is measured over a wider catalog when one is given, because unlike
    # accuracy it does not need an answer key and is far more meaningful at scale.
    if input_path is not None:
        with console.status(f"[bold]measuring compliance over {input_path.name}..."):
            wide = run_enrich(input_path, limit=limit, model=model)
    else:
        wide = scored

    schemas = dict(load_all())
    schemas[GENERIC_CATEGORY_ID] = generic_schema()

    limits = check_limits(wide.rows)
    vocabulary = check_vocabulary(wide.records, schemas)

    console.print()
    console.print(f"[dim]compliance measured over {len(wide.rows)} product(s)[/dim]")
    console.print()
    for line in format_report(accuracy, limits, vocabulary):
        if line.startswith("  !"):
            console.print(f"[yellow]{line}[/yellow]")
        elif line and not line.startswith(" "):
            console.print(f"[bold]{line}[/bold]")
        else:
            console.print(line)
